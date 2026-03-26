from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from typing import List
from app.database import get_db
from app.models import db_models
from app import schemas

router = APIRouter()

from fastapi import UploadFile, File, Query
import openpyxl
import io

# --- IMPORT MACHINES FROM EXCEL ---
@router.post("/machines/import")
def import_machines(
    file: UploadFile = File(...),
    mode: str = Query("append"),
    db: Session = Depends(get_db),
):
    contents = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    if mode == "replace":
        db.query(db_models.Machine).delete()
        db.commit()

    created = []
    skipped = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        code = data.get("id") or data.get("code")

        if not code:
            continue

        existing = db.query(db_models.Machine).filter(db_models.Machine.code == str(code)).first()
        if existing:
            skipped.append(code)
            continue

        db_machine = db_models.Machine(
            code=str(code),
            name=str(data.get("name") or code),
            group=str(data.get("group", "medium")),
            tonnage=int(data.get("tonnage", 0)),
            hours_per_day=float(data.get("hours_per_day", 24.0)),
            efficiency=float(data.get("efficiency", 1.0)),
            status=str(data.get("status", "available")),
        )
        db.add(db_machine)
        created.append(code)

    db.commit()
    return {"created": len(created), "skipped": len(skipped), "codes": created}


# --- IMPORT MOLDS FROM EXCEL ---
@router.post("/molds/import")
def import_molds(
    file: UploadFile = File(...),
    mode: str = Query("append"),
    db: Session = Depends(get_db),
):
    contents = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]

    if mode == "replace":
        db.query(db_models.Mold).delete()
        db.commit()

    created = []
    skipped = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        
        # Accept 'id' or 'code' from Excel
        code = data.get("id") or data.get("code")
        if not code or str(code).lower() == "none":
            continue
            
        code = str(code).strip()

        existing = db.query(db_models.Mold).filter(db_models.Mold.code == code).first()
        if existing:
            skipped.append(code)
            continue

        db_mold = db_models.Mold(
            code=code,
            name=str(data.get("name") or code),
            group=str(data.get("group", "medium")).strip(),
            tonnage=int(data.get("tonnage", 0)),
            component_id=str(data.get("component_id")).strip() if data.get("component_id") else None,
        )
        db.add(db_mold)
        created.append(code)

    db.commit()
    return {"created": len(created), "skipped": len(skipped), "codes": created}


# --- IMPORT COMPONENTS FROM EXCEL ---
@router.post("/plans/{plan_id}/components/import")
def import_components(
    plan_id: str,
    file: UploadFile = File(...),
    mode: str = Query("append"),
    db: Session = Depends(get_db),
):
    db_plan = db.query(db_models.Plan).filter(db_models.Plan.id == plan_id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    contents = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]

    if mode == "replace":
        db.query(db_models.Component).filter(db_models.Component.plan_id == plan_id).delete()
        db.commit()

    created = []
    skipped = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        
        # Accept 'id' or 'component_id' from Excel
        comp_id = data.get("id") or data.get("component_id")
        if not comp_id or str(comp_id).lower() == "none":
            continue
            
        comp_id = str(comp_id).strip()

        # Check if already exists in THIS plan
        existing = db.query(db_models.Component).filter(
            db_models.Component.plan_id == plan_id,
            db_models.Component.component_id == comp_id
        ).first()
        
        if existing:
            skipped.append(comp_id)
            continue

        # Normalize dependency_mode (Excel 'wait' -> DB 'wait_all')
        raw_mode = str(data.get("dependency_mode", "wait_all")).strip().lower()
        dependency_mode = "parallel" if raw_mode == "parallel" else "wait_all"

        # Handle column name variants for transfer time
        transfer_time = int(
            data.get("dependency_transfer_time_minutes") 
            or data.get("transfer_time_minutes") 
            or 0
        )

        # Handle prerequisites (nan, None, or comma-separated string)
        raw_prereq = data.get("prerequisites")
        if not raw_prereq or str(raw_prereq).lower() in ("nan", "none", ""):
            prerequisites = []
        elif isinstance(raw_prereq, list):
            prerequisites = raw_prereq
        else:
            prerequisites = [p.strip() for p in str(raw_prereq).split(",") if p.strip()]

        db_component = db_models.Component(
            plan_id=plan_id,
            component_id=comp_id,
            order_code=str(data.get("order_code")).strip() if data.get("order_code") else None,
            name=str(data.get("name") or comp_id),
            quantity=int(data.get("quantity", 0)),
            finished=int(data.get("finished", 0)),
            cycle_time_sec=float(data.get("cycle_time_sec", 0)),
            mold_id=str(data.get("mold_id", "")).strip(),
            color=str(data.get("color", "")).strip(),
            start_date=str(data.get("start_date")).strip() if data.get("start_date") else None,
            due_date=str(data.get("due_date", "")).strip(),
            lead_time_days=int(data.get("lead_time_days") or 2),
            dependency_mode=dependency_mode,
            dependency_transfer_time_minutes=transfer_time,
            prerequisites=prerequisites,
        )
        db.add(db_component)
        created.append(comp_id)

    db.commit()
    return {"created": len(created), "skipped": len(skipped), "component_ids": created}

# --- PLANS ---

@router.get("/plans", response_model=List[schemas.Plan])
def get_plans(db: Session = Depends(get_db)):
    return db.query(db_models.Plan).all()

@router.post("/plans", response_model=schemas.Plan)
def create_plan(plan: schemas.PlanCreate, db: Session = Depends(get_db)):
    db_plan = db_models.Plan(**plan.model_dump())
    db.add(db_plan)
    db.commit()
    db.refresh(db_plan)
    return db_plan

@router.get("/plans/{plan_id}", response_model=schemas.Plan)
def get_plan(plan_id: str, db: Session = Depends(get_db)):
    db_plan = db.query(db_models.Plan).filter(db_models.Plan.id == plan_id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    return db_plan

@router.delete("/plans/{plan_id}")
def delete_plan(plan_id: str, db: Session = Depends(get_db)):
    db_plan = db.query(db_models.Plan).filter(db_models.Plan.id == plan_id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    db.delete(db_plan)
    db.commit()
    return {"message": "Plan deleted"}

@router.patch("/plans/{plan_id}", response_model=schemas.Plan)
def update_plan(plan_id: str, plan: schemas.PlanCreate, db: Session = Depends(get_db)):
    db_plan = db.query(db_models.Plan).filter(db_models.Plan.id == plan_id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    for key, value in plan.model_dump().items():
        setattr(db_plan, key, value)
    db_plan.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(db_plan)
    return db_plan

# --- MACHINES ---

@router.get("/machines", response_model=List[schemas.Machine])
def get_machines(db: Session = Depends(get_db)):
    return db.query(db_models.Machine).all()

@router.post("/machines", response_model=schemas.Machine)
def create_machine(machine: schemas.MachineCreate, db: Session = Depends(get_db)):
    db_machine = db_models.Machine(**machine.model_dump())
    db.add(db_machine)
    db.commit()
    db.refresh(db_machine)
    return db_machine

@router.delete("/machines/{machine_id}")
def delete_machine(machine_id: str, db: Session = Depends(get_db)):
    db_machine = db.query(db_models.Machine).filter(db_models.Machine.id == machine_id).first()
    if not db_machine:
        raise HTTPException(status_code=404, detail="Machine not found")
    db.delete(db_machine)
    db.commit()
    return {"message": "Machine deleted"}

@router.patch("/machines/{machine_id}", response_model=schemas.Machine)
def update_machine(machine_id: str, machine: schemas.MachineCreate, db: Session = Depends(get_db)):
    db_machine = db.query(db_models.Machine).filter(db_models.Machine.id == machine_id).first()
    if not db_machine:
        raise HTTPException(status_code=404, detail="Machine not found")
    for key, value in machine.model_dump().items():
        setattr(db_machine, key, value)
    db.commit()
    db.refresh(db_machine)
    return db_machine

# --- MOLDS ---

@router.get("/molds", response_model=List[schemas.Mold])
def get_molds(db: Session = Depends(get_db)):
    return db.query(db_models.Mold).all()

@router.post("/molds", response_model=schemas.Mold)
def create_mold(mold: schemas.MoldCreate, db: Session = Depends(get_db)):
    db_mold = db_models.Mold(**mold.model_dump())
    db.add(db_mold)
    db.commit()
    db.refresh(db_mold)
    return db_mold

@router.delete("/molds/{mold_id}")
def delete_mold(mold_id: str, db: Session = Depends(get_db)):
    db_mold = db.query(db_models.Mold).filter(db_models.Mold.id == mold_id).first()
    if not db_mold:
        raise HTTPException(status_code=404, detail="Mold not found")
    db.delete(db_mold)
    db.commit()
    return {"message": "Mold deleted"}

@router.patch("/molds/{mold_id}", response_model=schemas.Mold)
def update_mold(mold_id: str, mold: schemas.MoldCreate, db: Session = Depends(get_db)):
    db_mold = db.query(db_models.Mold).filter(db_models.Mold.id == mold_id).first()
    if not db_mold:
        raise HTTPException(status_code=404, detail="Mold not found")
    for key, value in mold.model_dump().items():
        setattr(db_mold, key, value)
    db.commit()
    db.refresh(db_mold)
    return db_mold

# --- COMPONENTS ---

@router.get("/plans/{plan_id}/components", response_model=List[schemas.Component])
def get_components(plan_id: str, db: Session = Depends(get_db)):
    return db.query(db_models.Component).filter(db_models.Component.plan_id == plan_id).all()

@router.post("/plans/{plan_id}/components", response_model=schemas.Component)
def create_component(plan_id: str, component: schemas.ComponentCreate, db: Session = Depends(get_db)):
    db_component = db_models.Component(**component.model_dump(), plan_id=plan_id)
    db.add(db_component)
    db.commit()
    db.refresh(db_component)
    return db_component

@router.delete("/components/{component_id}")
def delete_component(component_id: str, db: Session = Depends(get_db)):
    db_component = db.query(db_models.Component).filter(db_models.Component.id == component_id).first()
    if not db_component:
        raise HTTPException(status_code=404, detail="Component not found")
    db.delete(db_component)
    db.commit()
    return {"message": "Component deleted"}

@router.patch("/components/{component_id}", response_model=schemas.Component)
def update_component(component_id: str, component: schemas.ComponentCreate, db: Session = Depends(get_db)):
    db_component = db.query(db_models.Component).filter(db_models.Component.id == component_id).first()
    if not db_component:
        raise HTTPException(status_code=404, detail="Component not found")
    for key, value in component.model_dump().items():
        setattr(db_component, key, value)
    db.commit()
    db.refresh(db_component)
    return db_component

from app.models.models import Machine as GAMachine, Mold as GAMold, ProductComponent as GAComponent
from app.services.ga_scheduler import ga_optimize_v2
from datetime import datetime, date, time

# --- RUNS ---

@router.post("/plans/{plan_id}/run", response_model=schemas.Run)
def run_plan(plan_id: str, db: Session = Depends(get_db)):
    # 1. Fetch Plan
    db_plan = db.query(db_models.Plan).filter(db_models.Plan.id == plan_id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    # 2. Fetch Global Machines & Molds
    db_machines = db.query(db_models.Machine).all()
    db_molds = db.query(db_models.Mold).all()

    # 3. Fetch Plan-Specific Components
    db_components = db.query(db_models.Component).filter(db_models.Component.plan_id == plan_id).all()

    if not db_machines or not db_components:
        raise HTTPException(status_code=400, detail="Plan must have machines and components to run")

    # 4. Convert DB Models to GA Dataclasses
    # (Mapping db_machine.code -> GAMachine.id as discussed)
    ga_machines = [
        GAMachine(
            id=m.code,
            name=m.name or m.code,
            group=m.group,
            tonnage=m.tonnage,
            hours_per_day=m.hours_per_day,
            efficiency=m.efficiency,
            status=m.status
        ) for m in db_machines
    ]

    ga_molds = [
        GAMold(
            id=m.code,
            name=m.name or m.code,
            group=m.group,
            tonnage=m.tonnage,
            component_id=m.component_id
        ) for m in db_molds
    ]

    ga_components = [
        GAComponent(
            id=c.component_id,
            name=c.name or c.component_id,
            quantity=c.quantity,
            finished=c.finished,
            cycle_time_sec=c.cycle_time_sec,
            mold_id=c.mold_id,
            color=c.color,
            start_date=date.fromisoformat(c.start_date) if c.start_date else None,
            due_date=date.fromisoformat(c.due_date) if c.due_date else date.fromisoformat(db_plan.current_date),
            dependency_mode=c.dependency_mode,
            dependency_transfer_time_minutes=c.dependency_transfer_time_minutes,
            prerequisites=c.prerequisites,
            order_code=c.order_code
        ) for c in db_components
    ]

    # 5. Run GA Scheduler
    try:
        # Convert plan strings to date/time objects
        current_date = date.fromisoformat(db_plan.current_date)
        start_time = time.fromisoformat(db_plan.start_time)

        result = ga_optimize_v2(
            components=ga_components,
            machines=ga_machines,
            molds=ga_molds,
            month_days=db_plan.month_days,
            mold_change_time_minutes=db_plan.mold_change_time_minutes,
            color_change_time_minutes=db_plan.color_change_time_minutes,
            current_date=current_date,
            start_time=start_time,
            pop_size=db_plan.pop_size,
            n_generations=db_plan.n_generations,
            mutation_rate=db_plan.mutation_rate,
        )

        # 6. Save Run Result to DB
        db_run = db_models.Run(
            plan_id=plan_id,
            run_name=f"Run {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            score=result.get("score"),
            unmet=result.get("unmet"),
            assignments=result.get("assignments"),
            status="completed"
        )
        db.add(db_run)
        db.commit()
        db.refresh(db_run)

        return db_run

    except Exception as e:
        # Log the error and save a failed run
        db_run = db_models.Run(
            plan_id=plan_id,
            run_name=f"Failed Run {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            status="failed",
            unmet={"error": str(e)}
        )
        db.add(db_run)
        db.commit()
        raise HTTPException(status_code=500, detail=f"Scheduling failed: {str(e)}")

# --- RUN HISTORY ---

@router.get("/plans/{plan_id}/runs", response_model=List[schemas.Run])
def get_plan_runs(plan_id: str, db: Session = Depends(get_db)):
    return db.query(db_models.Run).filter(db_models.Run.plan_id == plan_id).order_by(db_models.Run.run_at.desc()).all()

@router.get("/runs/{run_id}", response_model=schemas.Run)
def get_run(run_id: str, db: Session = Depends(get_db)):
    db_run = db.query(db_models.Run).filter(db_models.Run.id == run_id).first()
    if not db_run:
        raise HTTPException(status_code=404, detail="Run not found")
    return db_run