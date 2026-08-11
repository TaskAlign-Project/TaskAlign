from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Any, Iterable, Mapping, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_NAME = "Production Schedule"

# Fixed columns before the day grid.
HEADERS = [
    ("Item", 6),
    ("Code", 15),
    ("Description", 34),
    ("Production Ord#", 16),
    ("Q'ty", 10),
    ("Finished", 10),
    ("Diff", 10),
    ("Mold", 14),
    ("Machine(s)", 18),
]
FIRST_DAY_COL = len(HEADERS) + 1  # day 1 lands here

ROW_TITLE = 2
ROW_SUBTITLE = 3
ROW_TOTAL = 5
ROW_DAYNUM = 6
ROW_DATE = 7
ROW_HEADER = 8
ROW_FIRST_DATA = 9

FONT = "Arial"
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
DONE_FILL = PatternFill("solid", fgColor="E2EFDA")


def _get(obj: Any, key: str, default: Any = None) -> Any:
    """Assignments are dicts when read back from the Run table and objects when
    passed straight from the scheduler."""
    if isinstance(obj, Mapping):
        return obj.get(key, default)
    return getattr(obj, key, default)


def _split_component_id(component_id: str) -> tuple[str, str]:
    """component_id is "<material>-<order>"; the material is the client's Code."""
    text = str(component_id or "")
    material, _, order = text.partition("-")
    return (material or text), order


def build_schedule_workbook(
    plan: Any,
    assignments: Sequence[Any],
    components: Sequence[Any],
    run: Any = None,
    revision: str = "0",
) -> Workbook:
    """Render one run as a workbook. Returns an openpyxl Workbook."""
    month_days = int(getattr(plan, "month_days", 30) or 30)
    last_day_col = FIRST_DAY_COL + month_days - 1
    last_col_letter = get_column_letter(last_day_col)

    try:
        anchor = date.fromisoformat(str(plan.current_date))
    except (TypeError, ValueError):
        anchor = date.today()

    # produced[component_id][day] = pieces
    produced: dict[str, dict[int, int]] = defaultdict(lambda: defaultdict(int))
    machines_used: dict[str, set[str]] = defaultdict(set)
    molds_used: dict[str, set[str]] = defaultdict(set)

    for a in assignments or []:
        if _get(a, "task_type") != "PRODUCE":
            continue
        cid = _get(a, "component_id")
        qty = int(_get(a, "produced_qty") or 0)
        day = int(_get(a, "day") or 0)
        if not cid or qty <= 0 or day <= 0:
            continue
        produced[cid][day] += qty
        if _get(a, "machine_id"):
            machines_used[cid].add(str(_get(a, "machine_id")))
        if _get(a, "mold_id"):
            molds_used[cid].add(str(_get(a, "mold_id")))

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # ---- title block ------------------------------------------------------
    ws.cell(ROW_TITLE, 3, getattr(plan, "name", "Production Plan")).font = Font(
        name=FONT, size=14, bold=True
    )
    ws.cell(ROW_TITLE, last_day_col - 2, f"DATE : {date.today().strftime('%d/%m/%Y')}").font = Font(
        name=FONT, size=10
    )
    ws.cell(
        ROW_SUBTITLE, 3,
        f"PRODUCTION SCHEDULE FOR MONTH  {anchor.strftime('%B %Y').upper()}",
    ).font = Font(name=FONT, size=11, bold=True)
    ws.cell(ROW_SUBTITLE, last_day_col - 2, f"REVISION : {revision}").font = Font(name=FONT, size=10)

    if run is not None and _get(run, "run_name"):
        ws.cell(ROW_SUBTITLE, 1, str(_get(run, "run_name"))).font = Font(
            name=FONT, size=9, italic=True
        )

    # ---- day number and date rows ----------------------------------------
    ws.cell(ROW_TOTAL, len(HEADERS), "Total").font = Font(name=FONT, size=9, bold=True)
    ws.cell(ROW_DAYNUM, len(HEADERS), "Day").font = Font(name=FONT, size=9, bold=True)
    ws.cell(ROW_DATE, len(HEADERS), "Date").font = Font(name=FONT, size=9, bold=True)

    for offset in range(month_days):
        col = FIRST_DAY_COL + offset
        day_no = offset + 1
        day_date = anchor + timedelta(days=offset)

        c = ws.cell(ROW_DAYNUM, col, day_no)
        c.font = Font(name=FONT, size=9, bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

        c = ws.cell(ROW_DATE, col, day_date.strftime("%d/%m"))
        c.font = Font(name=FONT, size=8)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        if day_date.weekday() >= 5:  # visually mark weekends; still schedulable
            c.font = Font(name=FONT, size=8, color="C00000")

    # ---- header row -------------------------------------------------------
    for idx, (label, width) in enumerate(HEADERS, start=1):
        c = ws.cell(ROW_HEADER, idx, label)
        c.font = Font(name=FONT, size=9, bold=True)
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = width

    for offset in range(month_days):
        col = FIRST_DAY_COL + offset
        c = ws.cell(ROW_HEADER, col, offset + 1)
        c.font = Font(name=FONT, size=9, bold=True)
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 7

    # ---- data rows --------------------------------------------------------
    row = ROW_FIRST_DATA
    for item_no, comp in enumerate(components, start=1):
        cid = str(_get(comp, "component_id") or "")
        code, order = _split_component_id(cid)
        quantity = int(_get(comp, "quantity") or 0)
        finished = int(_get(comp, "finished") or 0)
        by_day = produced.get(cid, {})

        values = [
            item_no,
            code,
            _get(comp, "name") or "",
            _get(comp, "order_code") or order,
            quantity,
            finished,
            None,  # Diff, formula below
            ", ".join(sorted(molds_used.get(cid) or {str(_get(comp, "mold_id") or "")} - {""})),
            ", ".join(sorted(machines_used.get(cid, []))),
        ]
        for idx, value in enumerate(values, start=1):
            c = ws.cell(row, idx, value)
            c.font = Font(name=FONT, size=9)
            c.border = BORDER
            if idx in (1, 5, 6, 7):
                c.alignment = Alignment(horizontal="right")

        # Diff = ordered - already finished - scheduled this run. A formula, so
        # the sheet stays correct if someone edits a day cell by hand.
        ws.cell(row, 7).value = (
            f"=E{row}-F{row}-SUM({get_column_letter(FIRST_DAY_COL)}{row}:{last_col_letter}{row})"
        )

        for offset in range(month_days):
            col = FIRST_DAY_COL + offset
            qty = by_day.get(offset + 1)
            c = ws.cell(row, col, qty if qty else None)
            c.font = Font(name=FONT, size=9)
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")

        # Orders already fulfilled before this plan: shown, not scheduled.
        if quantity > 0 and finished >= quantity:
            for idx in range(1, len(HEADERS) + 1):
                ws.cell(row, idx).fill = DONE_FILL

        row += 1

    last_data_row = row - 1

    # ---- totals -----------------------------------------------------------
    if last_data_row >= ROW_FIRST_DATA:
        for offset in range(month_days):
            col = FIRST_DAY_COL + offset
            letter = get_column_letter(col)
            c = ws.cell(
                ROW_TOTAL, col,
                f"=SUM({letter}{ROW_FIRST_DATA}:{letter}{last_data_row})",
            )
            c.font = Font(name=FONT, size=9, bold=True)
            c.fill = TOTAL_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")

        for idx, letter in ((5, "E"), (6, "F"), (7, "G")):
            c = ws.cell(
                ROW_TOTAL, idx,
                f"=SUM({letter}{ROW_FIRST_DATA}:{letter}{last_data_row})",
            )
            c.font = Font(name=FONT, size=9, bold=True)
            c.fill = TOTAL_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal="right")

    # Keep the identifying columns and the header rows visible while scrolling
    # across 33 day columns.
    ws.freeze_panes = ws.cell(ROW_FIRST_DATA, FIRST_DAY_COL)
    ws.sheet_view.showGridLines = False
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_schedule_xlsx(
    plan: Any,
    assignments: Sequence[Any],
    components: Sequence[Any],
    run: Any = None,
    revision: str = "0",
) -> bytes:
    return workbook_to_bytes(
        build_schedule_workbook(plan, assignments, components, run=run, revision=revision)
    )


def suggested_filename(plan: Any, run: Any = None) -> str:
    name = str(getattr(plan, "name", "plan")).strip().replace(" ", "_") or "plan"
    stamp = datetime.now().strftime("%Y%m%d")
    return f"production_schedule_{name}_{stamp}.xlsx"