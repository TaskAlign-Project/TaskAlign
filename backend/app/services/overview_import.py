from __future__ import annotations

import io
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Iterator, Mapping, Optional, Sequence

import openpyxl

# --------------------------------------------------------------------------
# Sheet geometry
# --------------------------------------------------------------------------

OVERVIEW_SHEET = "Overview (3)"
OVERVIEW_HEADER_ROW = 1
OVERVIEW_FIRST_DATA_ROW = 3      # row 2 is an aggregate/units row
OVERVIEW_LAST_DATA_ROW = 443     # everything below is template padding

ZPPI_SHEET = "ZPPI010"
ZPPI_HEADER_ROW = 3
ZPPI_FIRST_DATA_ROW = 4

OV_MOLD_CODE = "Mold code"
OV_MATERIAL = "Material number"
OV_DESCRIPTION = "Material Description"
OV_GROUP = "Machine group"
OV_TONNAGE = "Size (Ton)"
OV_CYCLE = "ST. Time per piece (sec.)"
OV_ID_MACHINE = "ID Machine"

ZP_MATERIAL = "Material Code"
ZP_ORDER = "Order"
ZP_PLANNED = "Planned Q'ty"
ZP_PRODUCED = "Produced Q'ty"
ZP_START = "Start date"
ZP_DUE = "Due date"
ZP_DESCRIPTION = "Description"

DEFAULT_LEAD_TIME_DAYS = 2

# --------------------------------------------------------------------------
# Report
# --------------------------------------------------------------------------


@dataclass
class ImportReport:
    """Errors abort the import; warnings are advisory and shown in the dialog."""

    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    stats: dict[str, Any] = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        return not self.errors

    def error(self, msg: str) -> None:
        self.errors.append(msg)

    def warn(self, msg: str) -> None:
        self.warnings.append(msg)

    def as_dict(self) -> dict[str, Any]:
        return {"errors": self.errors, "warnings": self.warnings, "stats": self.stats}


# --------------------------------------------------------------------------
# Scalar normalisation
# --------------------------------------------------------------------------

GROUP_MAP = {
    "SMALL": "small",
    "MIDDLE": "medium",
    "MEDIUM": "medium",
    "BIG": "large",
    "LARGE": "large",
}


def normalize_group(value: Any) -> Optional[str]:
    """SMALL/MIDDLE/BIG -> small/medium/large. Unknown values return None.

    Returning None rather than silently defaulting matters: ga_scheduler
    compares mold.group to machine.group with ``!=``, so a wrong guess makes
    the component quietly unschedulable instead of loudly rejected.
    """
    if value is None:
        return None
    return GROUP_MAP.get(str(value).strip().upper())


def normalize_key(value: Any) -> str:
    """Material numbers arrive as int (Overview) or apostrophe-prefixed text
    (ZPPI, e.g. ``'300000006233``). Floats also pick up a trailing ``.0``."""
    if value is None:
        return ""
    text = str(value).strip().lstrip("'").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def to_int(value: Any, default: int = 0) -> int:
    """openpyxl returns None for blank cells, and ``int(None)`` raises."""
    if value is None or value == "":
        return default
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return default


def to_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def parse_date(value: Any) -> Optional[str]:
    """Accept DD.MM.YYYY (the ZPPI format), real datetimes, and ISO strings."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip().lstrip("'").strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


# --------------------------------------------------------------------------
# Colour extraction
# --------------------------------------------------------------------------

# Longest-first: "LIGHT GRAY" must win over "GRAY".
COLOR_TERMS: tuple[str, ...] = (
    "LIGHT VANILLA", "DARK VANILLA", "LIGHT GRAY", "LIGHT GREY",
    "DARK GRAY", "DARK GREY", "DARK GARY", "LIGHT PINK", "DARK PINK",
    "LIGHT BLUE", "DARK BLUE", "NEW WHITE", "NEW GRAY", "TURQUOISE",
    "VANILLA", "NATURAL", "MARINE", "VIOLET", "SILVER", "ORANGE",
    "PURPLE", "YELLOW", "WHITE", "BLACK", "GREEN", "BROWN", "CREAM",
    "GRAY", "GREY", "BLUE", "GOLD", "PINK", "RED",
)

# Typos and variants seen in the client's data, folded to one spelling so the
# GA does not charge a colour changeover between "GRAY" and "GARY".
COLOR_ALIASES = {
    "DARK GARY": "DARK GRAY",
    "GREY": "GRAY",
    "LIGHT GREY": "LIGHT GRAY",
    "DARK GREY": "DARK GRAY",
    "NEW WHITE": "WHITE",
    "NEW GRAY": "GRAY",
}

_COLOR_PATTERNS = (
    r"'([^']*)'",      # LEG KSH-D15/D18'BLUE'
    r"\(([^)]*)\)",    # HANDLE COVER KS-A18TTIND (GRAY)
    r"'([A-Z ]+)$",    # LEG KSH-206/211'GREEN   <- unterminated quote
)


def _canonical(term: str) -> str:
    return COLOR_ALIASES.get(term, term)


def extract_color(description: Any) -> str:
    """Best-effort colour from the material description.

    Roughly two thirds of rows state a colour; the rest are parts whose colour
    simply is not recorded. Those return "" on purpose -- ga_scheduler compares
    ``current_color != comp.color``, so all blank-colour parts match each other
    and incur no changeover, which is the behaviour we want.
    """
    if description is None:
        return ""
    text = str(description).upper()

    for pattern in _COLOR_PATTERNS:
        for candidate in re.findall(pattern, text):
            token = candidate.strip()
            if not token:
                continue
            for term in COLOR_TERMS:
                if token == term or token.startswith(term):
                    return _canonical(term)

    # Trailing "WH" suffix, e.g. LEG KSH-D18WH
    if re.search(r"WH$", text):
        return "WHITE"

    for term in COLOR_TERMS:
        if re.search(r"\b" + re.escape(term) + r"\b", text):
            return _canonical(term)

    return ""


# --------------------------------------------------------------------------
# Workbook reading
# --------------------------------------------------------------------------


def _open_sheet(contents: bytes, sheet_name: str, report: ImportReport):
    """data_only=True is required: these workbooks are formula-driven with
    cross-workbook links, so without it we would read '=CJ3*CD3' as a string."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surfaced to the user verbatim
        report.error(f"Could not open workbook: {exc}")
        return None

    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    report.warn(
        f'Sheet "{sheet_name}" not found (found: {", ".join(wb.sheetnames)}); '
        f'using "{wb.sheetnames[0]}".'
    )
    return wb[wb.sheetnames[0]]


def _iter_rows(
    ws,
    header_row: int,
    first_data_row: int,
    last_data_row: Optional[int],
    required: Sequence[str],
    report: ImportReport,
    label: str,
) -> Iterator[tuple[int, dict[str, Any]]]:
    """Yield (excel_row_number, {header: value}) for the data band."""
    headers: list[Optional[str]] = []
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        headers = [str(c).strip() if c is not None else None for c in row]
        break

    missing = [c for c in required if c not in headers]
    if missing:
        report.error(f"{label}: missing expected column(s): {', '.join(missing)}")
        return

    # ZPPI010 repeats "Start date", "Due date" and "Un" in later columns, where
    # the second copy holds a concatenated "material-date" string rather than a
    # date. dict(zip(...)) would silently keep the LAST occurrence, so map each
    # header to its first column index instead.
    index_of: dict[str, int] = {}
    duplicates: list[str] = []
    for idx, name in enumerate(headers):
        if name is None:
            continue
        if name in index_of:
            duplicates.append(name)
            continue
        index_of[name] = idx
    if duplicates:
        report.warn(
            f"{label}: duplicate column header(s) {', '.join(sorted(set(duplicates)))}; "
            f"using the leftmost of each."
        )

    for offset, row in enumerate(
        ws.iter_rows(min_row=first_data_row, max_row=last_data_row, values_only=True)
    ):
        yield first_data_row + offset, {
            name: (row[idx] if idx < len(row) else None) for name, idx in index_of.items()
        }


# --------------------------------------------------------------------------
# Overview
# --------------------------------------------------------------------------


@dataclass
class OverviewRow:
    material: str
    mold_code: str
    description: str
    group: Optional[str]
    tonnage: int
    cycle_time_sec: float
    color: str
    excel_row: int
    id_machine: str = ""
    # Filled in by build_molds() once it knows whether this row's (group, tonnage)
    # needed to be split into its own mold. Defaults to mold_code for callers
    # (tests, mainly) that construct a row without going through build_molds().
    resolved_mold_id: str = ""
    # Set by build_molds() when this row's (group, tonnage) matches no real
    # machine at all. build_components() drops the order instead of importing
    # a component that can never be scheduled.
    rejected_reason: str = ""


def read_overview(
    contents: bytes,
    report: ImportReport,
    last_data_row: Optional[int] = OVERVIEW_LAST_DATA_ROW,
) -> dict[str, OverviewRow]:
    """Return {material_number: OverviewRow} for the real data band."""
    ws = _open_sheet(contents, OVERVIEW_SHEET, report)
    if ws is None:
        return {}

    rows: dict[str, OverviewRow] = {}
    duplicates = 0
    no_group = 0
    no_cycle = 0

    for excel_row, data in _iter_rows(
        ws,
        OVERVIEW_HEADER_ROW,
        OVERVIEW_FIRST_DATA_ROW,
        last_data_row,
        [OV_MOLD_CODE, OV_MATERIAL, OV_DESCRIPTION, OV_GROUP, OV_TONNAGE, OV_CYCLE],
        report,
        "Overview",
    ):
        material = normalize_key(data.get(OV_MATERIAL))
        if not material:
            continue

        mold_code = str(data.get(OV_MOLD_CODE) or "").strip()
        if not mold_code:
            report.warn(f"Overview row {excel_row}: material {material} has no mold code; skipped.")
            continue

        group = normalize_group(data.get(OV_GROUP))
        if group is None:
            no_group += 1
            report.warn(
                f"Overview row {excel_row}: unrecognised machine group "
                f"{data.get(OV_GROUP)!r} for material {material}; skipped."
            )
            continue

        cycle = to_float(data.get(OV_CYCLE))
        if cycle <= 0:
            no_cycle += 1
            report.warn(
                f"Overview row {excel_row}: material {material} has no cycle time; skipped."
            )
            continue

        description = str(data.get(OV_DESCRIPTION) or "").strip()
        row = OverviewRow(
            material=material,
            mold_code=mold_code,
            description=description,
            group=group,
            tonnage=to_int(data.get(OV_TONNAGE)),
            cycle_time_sec=cycle,
            color=extract_color(description),
            excel_row=excel_row,
            # Optional: absent on sheets that don't have this column, in which
            # case build_molds() simply has nothing to cross-check against.
            id_machine=str(data.get(OV_ID_MACHINE) or "").strip(),
        )

        if material in rows:
            duplicates += 1
            report.warn(
                f"Overview row {excel_row}: material {material} already seen on row "
                f"{rows[material].excel_row}; keeping the first."
            )
            continue
        rows[material] = row

    report.stats["overview_materials"] = len(rows)
    report.stats["overview_duplicate_materials"] = duplicates
    report.stats["overview_skipped_no_group"] = no_group
    report.stats["overview_skipped_no_cycle"] = no_cycle
    report.stats["overview_colors_resolved"] = sum(1 for r in rows.values() if r.color)
    return rows


def build_molds(
    rows: Mapping[str, OverviewRow],
    report: ImportReport,
    machine_specs: Optional[Mapping[str, tuple[str, int]]] = None,
) -> list[dict[str, Any]]:
    """Collapse material rows into unique molds, one per (mold code, group, tonnage).

    A mold code serves several material numbers because the same part is produced
    in several colours -- those safely collapse into one mold. But the same mold
    code occasionally shows up against a *different* machine group/tonnage on
    other rows (e.g. logged against a 90T small press for some colours and a 160T
    medium press for another). Merging those by taking the majority group and the
    maximum tonnage used to produce an impossible mold -- e.g. group=small with
    tonnage=160, which no small machine satisfies -- silently making every
    component on that mold unschedulable. Instead, each distinct (group, tonnage)
    pair for a mold code becomes its own mold record, so every component still
    matches a real machine.

    ``machine_specs`` (machine code -> (real group, real tonnage), from the
    machines table) catches a related problem: Overview's own "Machine group"
    and "Size (Ton)" columns are free-typed and occasionally wrong for the
    machine named in "ID Machine" -- e.g. a 190T press logged as a 350T LARGE
    machine, so every material on it becomes "needs a 350T press", which
    nothing satisfies. Where a row's ID Machine is recognised, the machine's
    real group and tonnage replace whatever Overview typed for that row.

    When ID Machine is blank, unrecognised, or the resulting (group, tonnage)
    still matches no real machine of that group -- there is nothing to correct
    it to, so the row is rejected outright rather than turned into a mold no
    machine can ever run: no mold is created, its component(s) are dropped in
    build_components(), and one warning names exactly which mold and materials
    need fixing in the source file before the next import.

    component_id is deliberately left None: one mold maps to many materials, so
    no single value is correct, and ga_scheduler never reads the field.
    """
    if machine_specs:
        for row in rows.values():
            spec = machine_specs.get(row.id_machine)
            if spec is None:
                continue
            real_group, real_tonnage = spec
            if real_group != row.group or real_tonnage != row.tonnage:
                report.warn(
                    f"Overview row {row.excel_row}: material {row.material} labels "
                    f"machine {row.id_machine} as {row.group!r}/{row.tonnage}T, but "
                    f"the machines table says {real_group!r}/{real_tonnage}T; using "
                    f"the machines table."
                )
                row.group = real_group
                row.tonnage = real_tonnage

    max_tonnage_by_group: dict[str, int] = {}
    if machine_specs:
        for real_group, real_tonnage in machine_specs.values():
            if real_group is not None:
                max_tonnage_by_group[real_group] = max(
                    max_tonnage_by_group.get(real_group, 0), real_tonnage
                )

    rejected_by_variant: dict[tuple[str, Optional[str], int], list[OverviewRow]] = defaultdict(list)
    if max_tonnage_by_group:
        for row in rows.values():
            cap = max_tonnage_by_group.get(row.group, 0)
            if row.tonnage > cap:
                row.rejected_reason = (
                    f"no {row.group} machine can reach {row.tonnage}T "
                    f"(largest available {row.group} machine is {cap}T)"
                    if cap
                    else f"no {row.group} machine exists at all"
                )
                rejected_by_variant[(row.mold_code, row.group, row.tonnage)].append(row)

    for (mold_code, group, tonnage), bad_rows in rejected_by_variant.items():
        materials = ", ".join(sorted({r.material for r in bad_rows}))
        report.warn(
            f"Mold {mold_code}: needs a {group} machine rated {tonnage}T, but "
            f"{bad_rows[0].rejected_reason}. Not imported -- fix the group/tonnage "
            f"for material(s) {materials} in Overview and re-import."
        )

    variants: dict[tuple[str, Optional[str], int], list[OverviewRow]] = defaultdict(list)
    for row in rows.values():
        if row.rejected_reason:
            continue
        variants[(row.mold_code, row.group, row.tonnage)].append(row)

    rows_by_code: dict[str, list[tuple[str, Optional[str], int]]] = defaultdict(list)
    for key in variants:
        rows_by_code[key[0]].append(key)

    def variant_suffix(group: Optional[str], tonnage: int) -> str:
        # Both group and tonnage must be in the id: two variants can share a
        # tonnage with different groups (or vice versa), and either alone can
        # collide into the same "<code>-..." string for two distinct molds.
        return f"{group}-{tonnage}T"

    molds: list[dict[str, Any]] = []
    obsoleted_codes: set[str] = set()
    for code, keys in rows_by_code.items():
        split = len(keys) > 1
        # Deterministic order regardless of dict/materials iteration order.
        for _, group, tonnage in sorted(keys, key=lambda k: (k[2], k[1] or "")):
            variant_rows = variants[(code, group, tonnage)]
            mold_id = f"{code}-{variant_suffix(group, tonnage)}" if split else code
            for row in variant_rows:
                row.resolved_mold_id = mold_id
            molds.append({
                "code": mold_id,
                "name": mold_id,   # scheduler joins on name; keep them identical
                "group": group,
                "tonnage": tonnage,
                "component_id": None,
            })
        if split:
            variant_names = ", ".join(
                f"{code}-{variant_suffix(g, t)}" for _, g, t in sorted(keys, key=lambda k: (k[2], k[1] or ""))
            )
            report.warn(
                f"Mold {code}: found {len(keys)} different (group, tonnage) combinations "
                f"across materials; split into separate molds ({variant_names})."
            )
            obsoleted_codes.add(code)

    molds.sort(key=lambda m: m["code"])
    report.stats["molds"] = len(molds)
    # The bare code (e.g. "11C202AN") is superseded by its split variants and no
    # longer produced by this function. persist_molds() uses this to remove the
    # old merged row left over from before a split -- but only if nothing still
    # references it, so an older plan's components are never orphaned.
    report.stats["obsoleted_mold_codes"] = sorted(obsoleted_codes)
    return molds


# --------------------------------------------------------------------------
# ZPPI010
# --------------------------------------------------------------------------


@dataclass
class OrderRow:
    material: str
    order: str
    planned: int
    produced: int
    start_date: Optional[str]
    due_date: Optional[str]
    description: str
    excel_row: int


def read_orders(contents: bytes, report: ImportReport) -> list[OrderRow]:
    ws = _open_sheet(contents, ZPPI_SHEET, report)
    if ws is None:
        return []

    orders: list[OrderRow] = []
    seen: set[tuple[str, str]] = set()
    bad_dates = 0

    for excel_row, data in _iter_rows(
        ws,
        ZPPI_HEADER_ROW,
        ZPPI_FIRST_DATA_ROW,
        None,
        [ZP_MATERIAL, ZP_ORDER, ZP_PLANNED, ZP_PRODUCED, ZP_START, ZP_DUE],
        report,
        "ZPPI010",
    ):
        material = normalize_key(data.get(ZP_MATERIAL))
        if not material:
            continue  # the sheet is padded with thousands of blank rows

        order = normalize_key(data.get(ZP_ORDER))
        if not order:
            report.warn(f"ZPPI010 row {excel_row}: material {material} has no order number; skipped.")
            continue

        if (material, order) in seen:
            report.warn(f"ZPPI010 row {excel_row}: duplicate order {order} for {material}; skipped.")
            continue
        seen.add((material, order))

        start_date = parse_date(data.get(ZP_START))
        due_date = parse_date(data.get(ZP_DUE))
        if start_date is None or due_date is None:
            bad_dates += 1
            report.warn(
                f"ZPPI010 row {excel_row}: order {order} has an unreadable date "
                f"(start={data.get(ZP_START)!r}, due={data.get(ZP_DUE)!r})."
            )

        orders.append(
            OrderRow(
                material=material,
                order=order,
                planned=to_int(data.get(ZP_PLANNED)),
                produced=to_int(data.get(ZP_PRODUCED)),
                start_date=start_date,
                due_date=due_date,
                description=str(data.get(ZP_DESCRIPTION) or "").strip(),
                excel_row=excel_row,
            )
        )

    report.stats["orders"] = len(orders)
    report.stats["orders_bad_dates"] = bad_dates
    return orders


# --------------------------------------------------------------------------
# Join
# --------------------------------------------------------------------------


def build_components(
    overview: Mapping[str, OverviewRow],
    orders: Sequence[OrderRow],
    report: ImportReport,
    skip_completed: bool = False,
) -> list[dict[str, Any]]:
    """One component per production order.

    ``finished`` carries Produced Q'ty, so orders already fulfilled resolve to
    zero remaining work in check_unmet without needing to be filtered out.
    Set skip_completed=True to drop them from the GA entirely.
    """
    components: list[dict[str, Any]] = []
    unmatched: list[str] = []
    completed = 0
    zero_qty = 0
    rejected = 0

    for order in orders:
        source = overview.get(order.material)
        if source is None:
            unmatched.append(order.material)
            continue
        if source.rejected_reason:
            # Already warned about in build_molds(), which names the mold and
            # every affected material once rather than once per order here.
            rejected += 1
            continue

        remaining = max(order.planned - order.produced, 0)
        if remaining == 0:
            completed += 1
            if skip_completed:
                continue
        if order.planned <= 0:
            zero_qty += 1
            report.warn(
                f"Order {order.order} ({order.material}) has a planned quantity of "
                f"{order.planned}; imported but it will not be scheduled."
            )

        components.append(
            {
                "component_id": f"{order.material}-{order.order}",
                "order_code": order.order,
                "name": source.description or order.description or order.material,
                "quantity": order.planned,
                "finished": min(order.produced, order.planned),
                "cycle_time_sec": source.cycle_time_sec,
                # Mold identity: the scheduler joins on the mold name, which
                # build_molds() sets equal to this row's resolved_mold_id (the
                # mold code, or "<code>-<tonnage>T" if that code was split).
                "mold_id": source.resolved_mold_id or source.mold_code,
                "color": source.color,
                "start_date": order.start_date,
                "due_date": order.due_date,
                "lead_time_days": DEFAULT_LEAD_TIME_DAYS,
                # No dependency data exists in these exports; users add them later.
                "prerequisites": [],
                "dependency_mode": "wait_all",
                "dependency_transfer_time_minutes": 0,
            }
        )

    if unmatched:
        distinct = sorted(set(unmatched))
        preview = ", ".join(distinct[:10])
        more = f" (+{len(distinct) - 10} more)" if len(distinct) > 10 else ""
        report.warn(
            f"{len(unmatched)} order(s) across {len(distinct)} material(s) have no "
            f"Overview row and were skipped -- no mold or cycle time: {preview}{more}"
        )

    no_color = sum(1 for c in components if not c["color"])
    if no_color:
        report.warn(
            f"{no_color} of {len(components)} component(s) have no colour in the "
            f"material description. They are imported with an empty colour and "
            f"incur no colour changeover; set them manually if that is wrong."
        )

    report.stats["components"] = len(components)
    report.stats["orders_unmatched"] = len(unmatched)
    report.stats["orders_already_complete"] = completed
    report.stats["components_without_color"] = no_color
    report.stats["orders_zero_quantity"] = zero_qty
    report.stats["orders_rejected_infeasible"] = rejected
    return components


def estimate_capacity(
    components: Sequence[Mapping[str, Any]],
    molds: Sequence[Mapping[str, Any]],
) -> dict[str, float]:
    """Remaining machine-hours per machine group.

    The GA does not fail when the plan is over capacity -- it schedules what it
    can and returns the rest as unmet. Surfacing this at import time is what
    distinguishes "not enough machines" from "the scheduler is broken".
    """
    group_of = {m["name"]: m["group"] for m in molds}
    hours: dict[str, float] = defaultdict(float)
    for comp in components:
        remaining = max(int(comp["quantity"]) - int(comp["finished"]), 0)
        group = group_of.get(comp["mold_id"])
        if group and remaining:
            hours[group] += remaining * float(comp["cycle_time_sec"]) / 3600.0
    return dict(hours)


# --------------------------------------------------------------------------
# Entry point
# --------------------------------------------------------------------------


def import_from_client_files(
    overview_bytes: bytes,
    zppi_bytes: bytes,
    skip_completed: bool = False,
    overview_last_row: Optional[int] = OVERVIEW_LAST_DATA_ROW,
    machine_specs: Optional[Mapping[str, tuple[str, int]]] = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], ImportReport]:
    """Parse both workbooks and return (molds, components, report).

    ``machine_specs`` (machine code -> (real group, real tonnage)) lets
    build_molds() correct Overview rows whose typed group/tonnage disagree
    with the machine's real spec; see build_molds() for why that matters.
    Optional because plain parsing (e.g. in tests, or a dry preview with no
    machines loaded yet) should work without a database at hand.
    """
    report = ImportReport()

    overview = read_overview(overview_bytes, report, last_data_row=overview_last_row)
    if not report.ok:
        return [], [], report
    if not overview:
        report.error("No usable rows found in the Overview sheet.")
        return [], [], report

    molds = build_molds(overview, report, machine_specs=machine_specs)

    orders = read_orders(zppi_bytes, report)
    if not report.ok:
        return molds, [], report
    if not orders:
        report.error("No production orders found in the ZPPI010 sheet.")
        return molds, [], report

    components = build_components(overview, orders, report, skip_completed=skip_completed)
    report.stats["capacity_hours_by_group"] = {
        k: round(v, 1) for k, v in sorted(estimate_capacity(components, molds).items())
    }
    return molds, components, report